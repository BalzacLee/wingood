import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import requests
import json
import re
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, numbers
import threading
import pandas as pd
import os
from collections import defaultdict

# -------------------------- 全局变量与基础配置 --------------------------
# Wingood原有配置
BASE_URL = "https://city.wingoodcloud.com"
ACCOUNTS = [
    {"username": "xrqadmin", "password": "xrq@2024", "type": "wingood"},
    {"username": "liufu01", "password": "111111", "type": "wingood"},
    {"username": "lccx", "password": "cx123456", "type": "wingood"},
    {"username": "xiaorui", "password": "xr$RFV5tgb", "type": "wingood"},
    {"username": "dqadmin", "password": "111111", "type": "wingood"},
    # 新增ppone账号（带type标识）
    {"username": "116605882", "password": "10216b345bbc10b85c376a972eeff7d1", "type": "ppone"}
]
log_widget = None
global_mon_stat = {}  # Wingood全局月租统计变量

# PPONE专属配置（从ppone.py迁移）
PPONE_CONFIG = {
    "LOGIN_URL": "https://web.otcp.cn/login",
    "REMEMBER_ME": "true",
    "PARK_CONFIG": {
        "id": 14572,
        "name": "东莞市利成停车场"
    },
    "LIMIT": 10,
    "PAGE_START": 1,
    "PROXIES": {"http": "http://127.0.0.1:7897", "https": "http://127.0.0.1:7897"},
    "COMMON_HEADERS": {
        "accept": "application/json, text/javascript, */*; q=0.01",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "connection": "keep-alive",
        "host": "web.otcp.cn",
        "origin": "https://park.otcp.cn",
        "referer": "https://park.otcp.cn/",
        "sec-ch-ua": "\"Microsoft Edge\";v=\"143\", \"Chromium\";v=\"143\", \"Not A(Brand\";v=\"24\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\"",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36 Edg/143.0.0.0"
    },
    "FUNCTION_CONFIG": {
        "charge_out": {
            "url": "https://web.otcp.cn/parkOutCharge/list",
            "sheet_name": "收费出场信息",
            "columns": [
                "序号", "所属停车场", "计费类型", "车牌", "入口通道", "入场时间",
                "出口通道", "出场时间", "停车时长", "总额", "现金", "电子支付"
            ]
        },
        "recharge": {
            "url": "https://web.otcp.cn//parkPlatePostpone/list",
            "sheet_name": "充值续费信息",
            "columns": [
                "所属停车场", "车牌号", "计费类型", "开始时间", "有效期止",
                "应收金额", "实收金额", "支付方式", "操作时间", "地址",
                "支付详情", "操作员"
            ]
        },
        "monthly_car": {
            "url": "https://web.otcp.cn//parkCar/list",
            "sheet_name": "存量月租车统计",
            "columns": ["所属停车场", "存量月租"],
            "params": {
                "status": 1,
                "isHide": 0,
                "garageIds": 0,
                "chargeType": 2
            }
        }
    }
}
# PPONE全局临时变量
ppone_start_date = ""
ppone_end_date = ""

# -------------------------- 通用日志函数（统一输出到GUI日志框） --------------------------
def log(msg):
    if log_widget:
        log_widget.insert(tk.END, f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {msg}\n")
        log_widget.see(tk.END)
    # 同时打印到控制台
    print(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {msg}", flush=True)

# -------------------------- Wingood原有核心函数 --------------------------
def wingood_login(username, password):
    session = requests.Session()
    headers_login = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{BASE_URL}/login.html",
        "Origin": BASE_URL,
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "Accept": "application/json, text/javascript, */*; q=0.01"
    }
    try:
        resp1 = session.post(f"{BASE_URL}/LoginUserName", data={"userName": username}, headers=headers_login)
        resp1.raise_for_status()
    except Exception as e:
        log(f"❌ {username} - 用户名校验失败：{str(e)}")
        return None, False

    try:
        resp2 = session.post(f"{BASE_URL}/Login", data={"userName": username, "password": password}, headers=headers_login)
        resp2.raise_for_status()
        result = resp2.json()
        if result.get("flag"):
            log(f"🎉 {username} - Wingood登录成功")
            return session, True
        else:
            log(f"❌ {username} - Wingood登录失败：{result.get('msg', '未知错误')}")
            return None, False
    except Exception as e:
        log(f"❌ {username} - Wingood登录请求失败：{str(e)}")
        return None, False

def wingood_query_order(session, start_date, end_date, username):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{BASE_URL}/html/system/charge-list.html",
        "Origin": BASE_URL,
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json, text/plain, */*"
    }
    params_json = json.dumps({
        "parkId": "", "carNo": "", "payTime": start_date, "stopTime": end_date,
        "enterTime_start": "", "enterTime_end": "", "outTime_start": "", "outTime_end": "",
        "totalAmount": "0", "queryType": "1"
    }, ensure_ascii=False)

    all_data = []
    page_no = 1
    page_size = 100

    try:
        while True:
            resp = session.post(
                f"{BASE_URL}/ajax/ajaxQueryOrderInfo",
                data={"paramsJson": params_json, "pageNo": page_no, "pageSize": page_size},
                headers=headers
            )
            resp.raise_for_status()
            result = resp.json()

            page_data = result.get('listStr', [])
            if not page_data:
                break

            all_data.extend(page_data)

            total_page = result.get('totalPage', 1)
            if page_no >= total_page:
                break

            page_no += 1

        log(f"✅ {username} - 临时车订单查询成功，共{len(all_data)}条")
        return all_data

    except Exception as e:
        log(f"❌ {username} - 临时车订单查询失败：{str(e)}")
        return []

def wingood_query_mon_recharge(session, start_term, val_term, username):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{BASE_URL}/html/system/monReCharge-list.html",
        "Origin": BASE_URL,
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json, text/plain, */*"
    }
    params_json = json.dumps({
        "parkId": "", "monUserName": "", "phone": "", "carNo": "",
        "startTerm": start_term, "valTerm": val_term
    }, ensure_ascii=False)
    try:
        resp = session.post(f"{BASE_URL}/ajax/ajaxQueryMonReCharge",
                           data={"paramsJson": params_json, "pageNo": 1, "pageSize": 100},
                           headers=headers)
        resp.raise_for_status()
        result = resp.json()
        recharge_data = result.get('listStr', [])
        log(f"✅ {username} - 月租车充值查询成功，共{len(recharge_data)}条")
        return recharge_data
    except Exception as e:
        log(f"❌ {username} - 月租车充值查询失败：{str(e)}")
        return []

def wingood_get_all_mon_car(session, username):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36 Edg/142.0.0.0",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": f"{BASE_URL}/html/system/carManage-list.html",
        "Origin": BASE_URL,
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json, text/plain, */*"
    }
    all_data = []
    page_no = 1
    while True:
        params_json = json.dumps({
            "parkId": "", "carNo": "", "userName": "", "phone": "", "homeAddress": ""
        }, ensure_ascii=False)
        try:
            resp = session.post(f"{BASE_URL}/ajax/ajaxQueryMonCar",
                               data={"paramsJson": params_json, "pageNo": page_no, "pageSize": 100},
                               headers=headers)
            resp.raise_for_status()
            result = resp.json()
            if not result.get("flag"):
                break
            page_data = result.get("listStr", [])
            if not page_data:
                break
            all_data.extend(page_data)
            total_page = result.get("totalPage", 1)
            if page_no >= total_page:
                break
            page_no += 1
        except Exception as e:
            log(f"❌ {username} - 获取月租车数据第{page_no}页失败：{str(e)}")
            break
    log(f"✅ {username} - 月租车数据获取完成，共{len(all_data)}条")
    return all_data

def wingood_stat_mon_car(all_mon_car, username):
    global global_mon_stat
    park_stats = {}
    now = datetime.now()
    for car in all_mon_car:
        phone = car.get("phone", "").strip()
        clean_phone = re.sub(r"[\s\-\+()]", "", phone)
        if not re.match(r"^[1-9]\d{10}$", clean_phone):
            continue

        val_term_str = car.get("valTerm")
        if not val_term_str:
            continue
        try:
            if len(val_term_str) > 10:
                val_term = datetime.strptime(val_term_str, "%Y-%m-%d %H:%M:%S")
            else:
                val_term = datetime.strptime(val_term_str, "%Y-%m-%d")
        except ValueError:
            continue

        park_name = car.get("parkName", "未知停车场")
        charge_name = car.get("monChargeName", "默认规则") or "默认规则"

        if park_name not in park_stats:
            park_stats[park_name] = {
                "valid": 0,
                "expired_7d": 0,
                "rules": {}
            }

        if val_term >= now:
            park_stats[park_name]["valid"] += 1
            if charge_name not in park_stats[park_name]["rules"]:
                park_stats[park_name]["rules"][charge_name] = 0
            park_stats[park_name]["rules"][charge_name] += 1
        elif (now - val_term).days <= 7:
            park_stats[park_name]["expired_7d"] += 1

    for park_name, stats in park_stats.items():
        if park_name not in global_mon_stat:
            global_mon_stat[park_name] = {"valid": 0, "expired_7d": 0, "rules": {}}
        global_mon_stat[park_name]["valid"] += stats["valid"]
        global_mon_stat[park_name]["expired_7d"] += stats["expired_7d"]
        for rule, count in stats["rules"].items():
            if rule not in global_mon_stat[park_name]["rules"]:
                global_mon_stat[park_name]["rules"][rule] = 0
            global_mon_stat[park_name]["rules"][rule] += count

    log(f"✅ {username} - 月租车统计完成")
    return park_stats

# -------------------------- PPONE核心函数（从ppone.py迁移） --------------------------
def ppone_validate_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def ppone_charge_type_mapping_charge_out(charge_type):
    mapping = {1: "临时收费", 0: "免费", 2: "月租收费"}
    return mapping.get(charge_type, f"未知类型({charge_type})")

def ppone_charge_type_mapping_recharge(charge_type):
    mapping = {2: "月租", 1: "临时收费", 0: "免费", None: "未知"}
    return mapping.get(charge_type, f"未知类型({charge_type})")

def ppone_pay_type_mapping(pay_type):
    mapping = {1: "现金", 2: "电子支付", 3: "充正", None: "未知"}
    return mapping.get(pay_type, f"未知支付方式({pay_type})")

def ppone_calculate_monthly_car_duplicate(raw_data):
    group_dict = defaultdict(int)
    total_count = len(raw_data)

    for item in raw_data:
        phone = item.get("phone", "").strip()
        begin_date = item.get("beginDate", "")
        end_date = item.get("endDate", "")
        if phone:
            key = (phone, begin_date, end_date)
            group_dict[key] += 1

    duplicate_count = sum([count - 1 for count in group_dict.values() if count > 1])
    stock_count = total_count - duplicate_count

    log(f"📊 PPONE月租车统计：总条数{total_count} | 重复数{duplicate_count} | 存量月租{stock_count}")
    return total_count, duplicate_count, stock_count

def ppone_login(username, password):
    session = requests.Session()
    requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

    login_data = {
        "password": password,
        "username": username,
        "rememberMe": PPONE_CONFIG["REMEMBER_ME"]
    }
    try:
        resp = session.post(
            url=PPONE_CONFIG["LOGIN_URL"],
            headers=PPONE_CONFIG["COMMON_HEADERS"],
            data=login_data,
            verify=False,
            timeout=10,
            proxies=PPONE_CONFIG["PROXIES"]
        )
        resp.raise_for_status()
        login_result = resp.json()

        if login_result.get("code") == 0 and login_result.get("msg") == "SUCCESS":
            token = login_result.get("data")
            log(f"✅ {username} - PPONE登录成功，获取到token")
            return session, token
        else:
            log(f"❌ {username} - PPONE登录失败：{login_result.get('msg')}")
            return None, None
    except Exception as e:
        log(f"❌ {username} - PPONE登录异常：{str(e)}")
        return None, None

def ppone_query_all_pages(session, token, func_type):
    config = PPONE_CONFIG["FUNCTION_CONFIG"][func_type]
    time_start = f"{ppone_start_date} 00:00:00"
    time_end = f"{ppone_end_date} 00:00:00"
    log(f"\n📅 【PPONE-{func_type}】查询时间范围：{time_start} 至 {time_end}")

    all_data = []
    current_page = PPONE_CONFIG["PAGE_START"]

    base_params = {
        "page": current_page,
        "limit": PPONE_CONFIG["LIMIT"],
        "parkId": PPONE_CONFIG["PARK_CONFIG"]["id"],
        "token": token
    }

    if func_type == "charge_out":
        base_params.update({
            "timeStart": time_start,
            "timeEnd": time_end,
            "plate": "",
            "param1": 2,
            "admin": "",
            "chargeType": "",
            "outLane": "",
            "inLane": "",
            "payType": ""
        })
    elif func_type == "recharge":
        base_params.update({
            "timeStart": time_start,
            "timeEnd": time_end,
            "bindPlates": "",
            "address": "",
            "payType": "",
            "chargeType": "",
            "admin": ""
        })
    elif func_type == "monthly_car":
        base_params.update(config.get("params", {}))
        base_params.update({
            "plate": "",
            "name": "",
            "address": "",
            "cardNo": "",
            "phone": ""
        })

    try:
        resp = session.get(
            url=config["url"],
            headers=PPONE_CONFIG["COMMON_HEADERS"],
            params=base_params,
            verify=False,
            timeout=15,
            proxies=PPONE_CONFIG["PROXIES"]
        )
        resp.raise_for_status()
        first_page_result = resp.json()

        if first_page_result.get("code") != 0:
            log(f"❌ 【PPONE-{func_type}】第{current_page}页查询失败：{first_page_result.get('msg')}")
            return [], 0

        first_page_data = first_page_result.get("data", [])
        total_count = first_page_result.get("count", 0)
        all_data.extend(first_page_data)

        log(f"✅ 【PPONE-{func_type}】第{current_page}页查询成功，获取{len(first_page_data)}条，累计{len(all_data)}条")

        if func_type == "recharge":
            other = first_page_result.get("other", {})
            need_sum = other.get("needChargeSum", 0.0)
            actual_sum = other.get("actualChargeSum", 0.0)
            log(f"💰 【PPONE-{func_type}】汇总：应收总额{need_sum}元，实收总额{actual_sum}元")

        total_pages = (total_count + PPONE_CONFIG["LIMIT"] - 1) // PPONE_CONFIG["LIMIT"]
        log(f"📊 【PPONE-{func_type}】总条数：{total_count}，需查询{total_pages}页")

        if total_pages > 1:
            for current_page in range(PPONE_CONFIG["PAGE_START"] + 1, total_pages + 1):
                base_params["page"] = current_page
                try:
                    page_resp = session.get(
                        url=config["url"],
                        headers=PPONE_CONFIG["COMMON_HEADERS"],
                        params=base_params,
                        verify=False,
                        timeout=15,
                        proxies=PPONE_CONFIG["PROXIES"]
                    )
                    page_resp.raise_for_status()
                    page_result = page_resp.json()

                    if page_result.get("code") == 0:
                        page_data = page_result.get("data", [])
                        all_data.extend(page_data)
                        log(f"✅ 【PPONE-{func_type}】第{current_page}页查询成功，累计{len(all_data)}条")
                    else:
                        log(f"⚠️ 【PPONE-{func_type}】第{current_page}页查询失败：{page_result.get('msg')}，跳过该页")

                except Exception as e:
                    log(f"⚠️ 【PPONE-{func_type}】第{current_page}页查询异常：{str(e)}，跳过该页")

    except Exception as e:
        log(f"❌ 【PPONE-{func_type}】分页查询异常：{str(e)}")
        return [], 0

    log(f"✅ 【PPONE-{func_type}】所有分页查询完成！实际获取{len(all_data)}条")
    return all_data, total_count

def ppone_format_charge_out_data(raw_data):
    excel_data = []
    for idx, item in enumerate(raw_data, start=1):
        row = {
            "序号": idx,
            "所属停车场": PPONE_CONFIG["PARK_CONFIG"]["name"],
            "计费类型": ppone_charge_type_mapping_charge_out(item.get("chargeType", 0)),
            "车牌": item.get("plate", "未知车牌"),
            "入口通道": item.get("inLaneName", ""),
            "入场时间": item.get("inTime", ""),
            "出口通道": item.get("outLaneName", ""),
            "出场时间": item.get("outTime", ""),
            "停车时长": item.get("stopTimeStr", ""),
            "总额": float(item.get("needCharge", 0.0)),
            "现金": float(item.get("actualCharge", 0.0)),
            "电子支付": float(item.get("prepay", 0.0))
        }
        excel_data.append(row)
    return excel_data

def ppone_format_recharge_data(raw_data):
    excel_data = []
    for idx, item in enumerate(raw_data, start=1):
        row = {
            "所属停车场": PPONE_CONFIG["PARK_CONFIG"]["name"],
            "车牌号": item.get("plate", ""),
            "计费类型": ppone_charge_type_mapping_recharge(item.get("chargeType")),
            "开始时间": item.get("beginDate", ""),
            "有效期止": item.get("endDate", ""),
            "应收金额": float(item.get("needCharge", 0.0)),
            "实收金额": float(item.get("actualCharge", 0.0)),
            "支付方式": ppone_pay_type_mapping(item.get("payType")),
            "操作时间": item.get("createTime", ""),
            "地址": item.get("address", ""),
            "支付详情": item.get("remark", ""),
            "操作员": item.get("admin", "")
        }
        excel_data.append(row)
    return excel_data

def ppone_format_monthly_car_data(raw_data):
    total_count, duplicate_count, stock_count = ppone_calculate_monthly_car_duplicate(raw_data)
    excel_data = [{
        "所属停车场": PPONE_CONFIG["PARK_CONFIG"]["name"],
        "存量月租": stock_count
    }]
    return excel_data

def ppone_export_combined_excel(charge_out_data, recharge_data, monthly_car_data,
                                charge_out_count, recharge_count, monthly_car_count):
    if not charge_out_data and not recharge_data and not monthly_car_data:
        log("❌ PPONE无任何数据可导出！")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{PPONE_CONFIG['PARK_CONFIG']['name']}_综合查询_{ppone_start_date}_至_{ppone_end_date}_{timestamp}.xlsx"
    file_path = os.path.join(os.getcwd(), filename)

    try:
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            if charge_out_data:
                df_charge_out = pd.DataFrame(charge_out_data, columns=PPONE_CONFIG["FUNCTION_CONFIG"]["charge_out"]["columns"])
                df_charge_out.to_excel(writer, sheet_name=PPONE_CONFIG["FUNCTION_CONFIG"]["charge_out"]["sheet_name"], index=False)
                log(f"✅ PPONE写入工作表：{PPONE_CONFIG['FUNCTION_CONFIG']['charge_out']['sheet_name']}（{len(charge_out_data)}条）")

            if recharge_data:
                df_recharge = pd.DataFrame(recharge_data, columns=PPONE_CONFIG["FUNCTION_CONFIG"]["recharge"]["columns"])
                df_recharge.to_excel(writer, sheet_name=PPONE_CONFIG["FUNCTION_CONFIG"]["recharge"]["sheet_name"], index=False)
                log(f"✅ PPONE写入工作表：{PPONE_CONFIG['FUNCTION_CONFIG']['recharge']['sheet_name']}（{len(recharge_data)}条）")

            if monthly_car_data:
                df_monthly_car = pd.DataFrame(monthly_car_data, columns=PPONE_CONFIG["FUNCTION_CONFIG"]["monthly_car"]["columns"])
                df_monthly_car.to_excel(writer, sheet_name=PPONE_CONFIG["FUNCTION_CONFIG"]["monthly_car"]["sheet_name"], index=False)
                log(f"✅ PPONE写入工作表：{PPONE_CONFIG['FUNCTION_CONFIG']['monthly_car']['sheet_name']}（统计数据）")

        log(f"\n🎉 PPONE综合Excel导出成功！文件路径：{file_path}")
        log(f"📊 PPONE统计：收费出场{charge_out_count}条 | 充值续费{recharge_count}条 | 月租车总条数{monthly_car_count}条")
    except Exception as e:
        log(f"❌ PPONE Excel导出失败：{str(e)}")

def ppone_run_query(username, password, start_date, end_date):
    global ppone_start_date, ppone_end_date
    ppone_start_date = start_date
    ppone_end_date = end_date

    # PPONE登录
    login_session, token = ppone_login(username, password)
    if not login_session or not token:
        log(f"❌ {username} - PPONE登录失败，终止查询")
        return

    # 查询收费出场数据
    log("\n" + "-" * 50 + "【PPONE开始查询收费出场数据】" + "-" * 50)
    charge_out_raw, charge_out_count = ppone_query_all_pages(login_session, token, "charge_out")
    charge_out_formatted = ppone_format_charge_out_data(charge_out_raw)

    # 查询充值数据
    log("\n" + "-" * 50 + "【PPONE开始查询充值续费数据】" + "-" * 50)
    recharge_raw, recharge_count = ppone_query_all_pages(login_session, token, "recharge")
    recharge_formatted = ppone_format_recharge_data(recharge_raw)

    # 查询月租车数据
    log("\n" + "-" * 50 + "【PPONE开始查询有效期内月租车数据】" + "-" * 50)
    monthly_car_raw, monthly_car_count = ppone_query_all_pages(login_session, token, "monthly_car")
    monthly_car_formatted = ppone_format_monthly_car_data(monthly_car_raw)

    # 导出Excel
    log("\n" + "-" * 50 + "【PPONE开始导出综合Excel】" + "-" * 50)
    ppone_export_combined_excel(charge_out_formatted, recharge_formatted, monthly_car_formatted,
                                charge_out_count, recharge_count, monthly_car_count)

    log(f"\n✅ {username} - PPONE所有操作完成！")

# -------------------------- 统一导出函数（兼容Wingood和PPONE） --------------------------
def export_excel(all_results, filename):
    global global_mon_stat
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Wingood临时车订单工作表
    if all_results.get("order"):
        ws_order = wb.create_sheet(title="临时车订单（整合）")
        order_headers = [
            "编号", "停车场名称", "车牌号码", "入场时间", "入口车道",
            "出场时间", "出口车道", "支付金额", "支付订单号", "支付类型名称",
            "支付时间", "应付金额"
        ]
        order_fields = [
            "orderNo", "parkName", "carNo", "enterTime", "enterGateName",
            "outTime", "outGateName", "payMoney", "payOrderNo", "payTypeName",
            "payTime", "totalAmount"
        ]
        ws_order.append(order_headers)
        for item in all_results["order"]:
            row = [item.get(field, "") for field in order_fields]
            ws_order.append(row)
        for cell in ws_order[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws_order.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            ws_order.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    # 2. Wingood月租车充值工作表
    if all_results.get("recharge"):
        ws_recharge = wb.create_sheet(title="月租车充值（整合）")
        recharge_headers = [
            "车牌号", "停车场名称", "用户名称", "电话号码", "支付类型",
            "充值金额", "支付订单号", "支付时间"
        ]
        recharge_fields = [
            "carNo", "parkName", "monUserId", "phone", "chargeWay",
            "money", "payOrderNo", "payTime"
        ]
        ws_recharge.append(recharge_headers)
        for idx, item in enumerate(all_results["recharge"], start=2):
            row = []
            for field in recharge_fields:
                value = item.get(field, "")
                if field == "money" and value:
                    try:
                        value = float(value)
                    except ValueError:
                        value = 0
                row.append(value)
            ws_recharge.append(row)
            ws_recharge[f"F{idx}"].number_format = numbers.FORMAT_NUMBER_00
        for cell in ws_recharge[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws_recharge.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            ws_recharge.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    # 3. Wingood月租统计汇总
    if global_mon_stat:
        ws_stat = wb.create_sheet(title="月租统计汇总")
        all_rules = set()
        for park_data in global_mon_stat.values():
            all_rules.update(park_data["rules"].keys())
        all_rules = list(all_rules)
        stat_headers = ["停车场名称", "正常总数X（有效期内）", "7天内过期总数Y"] + all_rules
        ws_stat.append(stat_headers)

        for park_name, stats in global_mon_stat.items():
            row = [
                park_name,
                stats["valid"],
                stats["expired_7d"]
            ]
            for rule in all_rules:
                row.append(stats["rules"].get(rule, 0))
            ws_stat.append(row)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for cell in ws_stat[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row in ws_stat.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws_stat.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            ws_stat.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    wb.save(filename)
    log(f"📊 Wingood整合数据导出完成：{filename}")
    global_mon_stat = {}

# -------------------------- 主任务函数（分支处理Wingood/PPONE） --------------------------
def run_query(selected_accounts, start_date, end_date, func_order, func_recharge, func_mon_stat):
    global global_mon_stat
    # Wingood数据容器
    wingood_all_results = {
        "order": [],
        "recharge": []
    }
    global_mon_stat = {}

    # 遍历选中的账号，分支处理
    for account in selected_accounts:
        username = account["username"]
        password = account["password"]
        account_type = account.get("type", "wingood")

        # 分支1：PPONE账号（116605882）
        if account_type == "ppone":
            log(f"\n🚀 开始处理PPONE账号：{username}")
            ppone_run_query(username, password, start_date, end_date)
            continue

        # 分支2：Wingood账号
        log(f"\n🚀 开始处理Wingood账号：{username}")
        session, login_ok = wingood_login(username, password)
        if not login_ok:
            continue

        # Wingood临时车订单
        if func_order:
            order_data = wingood_query_order(session, start_date, end_date, username)
            if order_data:
                wingood_all_results["order"].extend(order_data)
                log(f"✅ {username} - 临时车订单已整合，累计{len(wingood_all_results['order'])}条")

        # Wingood月租车充值
        if func_recharge:
            recharge_data = wingood_query_mon_recharge(session, start_date, end_date, username)
            if recharge_data:
                wingood_all_results["recharge"].extend(recharge_data)
                log(f"✅ {username} - 月租车充值已整合，累计{len(wingood_all_results['recharge'])}条")

        # Wingood月租车统计
        if func_mon_stat:
            mon_data = wingood_get_all_mon_car(session, username)
            if mon_data:
                wingood_stat_mon_car(mon_data, username)

    # 导出Wingood整合数据（仅当有Wingood账号数据时）
    has_wingood_data = any([acc.get("type") == "wingood" for acc in selected_accounts])
    if has_wingood_data and (wingood_all_results["order"] or wingood_all_results["recharge"] or global_mon_stat):
        wingood_filename = f"Wingood整合数据统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        export_excel(wingood_all_results, wingood_filename)
    elif has_wingood_data:
        log("❌ Wingood无数据可导出")

    messagebox.showinfo("完成", "所有选中账号的查询任务执行完毕！")

# -------------------------- GUI界面（兼容新增账号类型） --------------------------
def create_gui():
    global log_widget
    root = tk.Tk()
    root.title("停车场运营统计工具（兼容Wingood/PPONE）")
    root.geometry("900x700")

    # 账号配置
    frame_account = ttk.LabelFrame(root, text="账号配置")
    frame_account.pack(fill=tk.X, padx=10, pady=5)

    listbox_accounts = tk.Listbox(frame_account, selectmode=tk.MULTIPLE, height=6)
    listbox_accounts.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
    for idx, acc in enumerate(ACCOUNTS):
        # 显示账号类型标识
        display_text = f"{acc['username']} ({acc['type']})"
        listbox_accounts.insert(tk.END, display_text)
        listbox_accounts.select_set(idx)

    # 新增账号区域（默认新增Wingood类型，如需新增PPONE需手动改type）
    frame_account_edit = ttk.Frame(frame_account)
    frame_account_edit.pack(side=tk.RIGHT, padx=5, pady=5)
    ttk.Label(frame_account_edit, text="账号：").grid(row=0, column=0)
    entry_user = ttk.Entry(frame_account_edit, width=15)
    entry_user.grid(row=0, column=1)
    ttk.Label(frame_account_edit, text="密码：").grid(row=1, column=0)
    entry_pwd = ttk.Entry(frame_account_edit, width=15, show="*")
    entry_pwd.grid(row=1, column=1)
    ttk.Label(frame_account_edit, text="类型：").grid(row=2, column=0)
    entry_type = ttk.Entry(frame_account_edit, width=15)
    entry_type.grid(row=2, column=1)
    entry_type.insert(0, "wingood")  # 默认wingood，PPONE需手动输入ppone

    def add_account():
        user = entry_user.get().strip()
        pwd = entry_pwd.get().strip()
        acc_type = entry_type.get().strip().lower()
        if not user or not pwd:
            messagebox.showwarning("警告", "账号/密码不能为空！")
            return
        if user in [acc["username"] for acc in ACCOUNTS]:
            messagebox.showwarning("警告", "账号已存在！")
            return
        if acc_type not in ["wingood", "ppone"]:
            messagebox.showwarning("警告", "类型只能是wingood/ppone！")
            return
        ACCOUNTS.append({"username": user, "password": pwd, "type": acc_type})
        listbox_accounts.insert(tk.END, f"{user} ({acc_type})")
        listbox_accounts.select_set(listbox_accounts.size() - 1)
        entry_user.delete(0, tk.END)
        entry_pwd.delete(0, tk.END)
        entry_type.delete(0, tk.END)
        entry_type.insert(0, "wingood")
        log(f"✅ 添加账号：{user}（类型：{acc_type}）")

    ttk.Button(frame_account_edit, text="添加账号", command=add_account).grid(row=3, columnspan=2, pady=2)

    # 日期范围
    frame_date = ttk.LabelFrame(root, text="查询范围（YYYY-MM-DD）")
    frame_date.pack(fill=tk.X, padx=10, pady=5)
    ttk.Label(frame_date, text="开始：").grid(row=0, column=0, padx=5)
    entry_start = ttk.Entry(frame_date)
    entry_start.grid(row=0, column=1, padx=5)
    entry_start.insert(0, (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d"))
    ttk.Label(frame_date, text="结束：").grid(row=0, column=2, padx=5)
    entry_end = ttk.Entry(frame_date)
    entry_end.grid(row=0, column=3, padx=5)
    entry_end.insert(0, datetime.now().strftime("%Y-%m-%d"))

    # 功能选择（PPONE账号会忽略此选择，固定执行收费+充值+月租查询）
    frame_func = ttk.LabelFrame(root, text="查询功能（仅Wingood生效）")
    frame_func.pack(fill=tk.X, padx=10, pady=5)
    var_order = tk.BooleanVar(value=True)
    ttk.Checkbutton(frame_func, text="临时车订单", variable=var_order).grid(row=0, column=0, padx=10)
    var_recharge = tk.BooleanVar(value=True)
    ttk.Checkbutton(frame_func, text="月租车充值", variable=var_recharge).grid(row=0, column=1, padx=10)
    var_mon_stat = tk.BooleanVar(value=True)
    ttk.Checkbutton(frame_func, text="月租车统计", variable=var_mon_stat).grid(row=0, column=2, padx=10)

    # 执行按钮
    def on_run():
        selected_indices = listbox_accounts.curselection()
        if not selected_indices:
            messagebox.showwarning("警告", "请选择账号！")
            return
        # 映射选中的显示文本到实际账号对象
        selected_accounts = []
        for idx in selected_indices:
            display_text = listbox_accounts.get(idx)
            username = display_text.split(" (")[0]
            for acc in ACCOUNTS:
                if acc["username"] == username:
                    selected_accounts.append(acc)
                    break

        start = entry_start.get().strip()
        end = entry_end.get().strip()
        if not start or not end:
            messagebox.showwarning("警告", "日期不能为空！")
            return
        try:
            datetime.strptime(start, "%Y-%m-%d")
            datetime.strptime(end, "%Y-%m-%d")
        except ValueError:
            messagebox.showwarning("警告", "日期格式错误！")
            return

        log_widget.delete(1.0, tk.END)
        log("🚀 开始执行任务（兼容Wingood/PPONE）...")
        threading.Thread(target=run_query, args=(selected_accounts, start, end, var_order.get(), var_recharge.get(), var_mon_stat.get()), daemon=True).start()

    ttk.Button(root, text="开始查询并导出（兼容版）", command=on_run).pack(pady=10)

    # 日志框
    frame_log = ttk.LabelFrame(root, text="执行日志")
    frame_log.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    log_widget = scrolledtext.ScrolledText(frame_log, wrap=tk.WORD)
    log_widget.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    log("✅ 工具就绪（兼容Wingood/PPONE），可配置查询！")

    root.mainloop()

if __name__ == "__main__":
    # 检查依赖
    try:
        import pandas as pd
        from openpyxl import Workbook
    except ImportError:
        log("❌ 缺少依赖包，请执行：pip install requests pandas openpyxl")
        exit(1)
    create_gui()